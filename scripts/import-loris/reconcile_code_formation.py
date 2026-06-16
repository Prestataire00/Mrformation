#!/usr/bin/env python3
"""
Réconciliation migration espace formation via la clé pivot « Code formation ».

DRY-RUN PAR DÉFAUT — aucune écriture. Mesure les 3 jointures avant toute correction :
  Code formation ──┬─→ session  (Suivi de l'activité (1) : titre + date début + heures prévues)
                   ├─→ client   (Suivi de l'activité des clients (1) : Entreprise/Client → clients.company_name)
                   └─→ enrollment (Suivi de l'activité des stagaires (1) : Nom + Code formation)

Produit un rapport (console + JSON) : taux de match par jointure + listes des non-matchés.
Réutilise le même .env.local / REST / helpers que loris_import.py (service role, MR uniquement).

Usage :
  python3 scripts/import-loris/reconcile_code_formation.py            # dry-run (défaut)
  (l'application sélective sera un script séparé, après revue de ce rapport)
"""

import hashlib
import json
import re
import sys
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

# ── Config (alignée sur loris_import.py) ───────────────────────────────────
MR_ENTITY_ID = "f8acea54-71ab-4a22-8cf3-4e7170543bf1"
DOWNLOADS = Path.home() / "Downloads"
REPORT_PATH = Path(__file__).parent / "reconcile_report.json"

# Nouveaux fichiers porteurs de la clé « Code formation » (versions du client, suffixe (1))
F_SESSIONS = "Suivi de l'activité (1).xlsx"
F_CLIENTS = "Suivi de l'activité des clients (1).xlsx"
F_STAGIAIRES = "Suivi de l'activité des stagaires (1).xlsx"


# ── Env ────────────────────────────────────────────────────────────────────
def load_env():
    env = {}
    env_path = Path(__file__).parent.parent.parent / ".env.local"
    if not env_path.exists():
        sys.exit(f"❌ .env.local introuvable à {env_path}")
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env


ENV = load_env()
SUPABASE_URL = ENV.get("NEXT_PUBLIC_SUPABASE_URL", "").rstrip("/")
SERVICE_ROLE = ENV.get("SUPABASE_SERVICE_ROLE_KEY", "")
if not SUPABASE_URL or not SERVICE_ROLE:
    sys.exit("❌ NEXT_PUBLIC_SUPABASE_URL ou SUPABASE_SERVICE_ROLE_KEY manquant")


def rest_get(table, **params):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + urllib.parse.urlencode(params, quote_via=urllib.parse.quote)
    req = urllib.request.Request(
        url,
        headers={"apikey": SERVICE_ROLE, "Authorization": f"Bearer {SERVICE_ROLE}"},
        method="GET",
    )
    try:
        with urllib.request.urlopen(req) as r:
            raw = r.read()
            return json.loads(raw) if raw else []
    except urllib.error.HTTPError as e:
        sys.exit(f"❌ REST {table}: {e.code} {e.read().decode(errors='replace')[:300]}")


def rest_get_all(table, **params):
    """Pagination par range (PostgREST limite à 1000 par défaut)."""
    out = []
    step = 1000
    start = 0
    while True:
        url = f"{SUPABASE_URL}/rest/v1/{table}"
        if params:
            url += "?" + urllib.parse.urlencode(params, quote_via=urllib.parse.quote)
        req = urllib.request.Request(
            url,
            headers={
                "apikey": SERVICE_ROLE,
                "Authorization": f"Bearer {SERVICE_ROLE}",
                "Range-Unit": "items",
                "Range": f"{start}-{start + step - 1}",
            },
            method="GET",
        )
        try:
            with urllib.request.urlopen(req) as r:
                raw = r.read()
                batch = json.loads(raw) if raw else []
        except urllib.error.HTTPError as e:
            sys.exit(f"❌ REST {table}: {e.code} {e.read().decode(errors='replace')[:300]}")
        out.extend(batch)
        if len(batch) < step:
            break
        start += step
    return out


# ── Helpers de normalisation (alignés loris_import.py) ─────────────────────
def norm(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s not in ("-", "—", "N/A", "NA") else None


def norm_name(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    return None


def stable_external_id(prefix, *parts):
    h = hashlib.sha256("|".join(str(p or "") for p in parts).encode()).hexdigest()[:16]
    return f"loris-{prefix}-{h}"


def read_xlsx(filename):
    path = DOWNLOADS / filename
    if not path.exists():
        sys.exit(f"❌ Fichier introuvable : {path}")
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [(str(h).strip() if h is not None else f"col_{i}") for i, h in enumerate(rows[0])]
    out = []
    for row in rows[1:]:
        d = {headers[i]: v for i, v in enumerate(row) if i < len(headers)}
        if any(v is not None and str(v).strip() for v in d.values()):
            out.append(d)
    return out


# ── Réconciliation ──────────────────────────────────────────────────────────
def main():
    print(f"Réconciliation Code formation — MR FORMATION ({MR_ENTITY_ID})")
    print(f"Supabase : {SUPABASE_URL}\n")

    # 1. État base
    db_sessions = rest_get_all("sessions", select="id,title,start_date,loris_external_id",
                               entity_id=f"eq.{MR_ENTITY_ID}")
    db_learners = rest_get_all("learners", select="id,first_name,last_name,client_id,loris_external_id",
                               entity_id=f"eq.{MR_ENTITY_ID}")
    db_clients = rest_get_all("clients", select="id,company_name",
                              entity_id=f"eq.{MR_ENTITY_ID}")
    print(f"Base : {len(db_sessions)} sessions, {len(db_learners)} learners, {len(db_clients)} clients\n")

    # Index base
    sess_by_extid = {s["loris_external_id"]: s for s in db_sessions if s.get("loris_external_id")}
    sess_by_title_date = {}
    sess_by_title = defaultdict(list)
    for s in db_sessions:
        t = norm_name(s.get("title"))
        d = (s.get("start_date") or "")[:10]
        sess_by_title_date[(t, d)] = s
        sess_by_title[t].append(s)

    client_by_name = defaultdict(list)
    for c in db_clients:
        client_by_name[norm_name(c.get("company_name"))].append(c["id"])

    # learners : index par "nom complet" normalisé (les 2 ordres), repère les homonymes
    learner_by_name = defaultdict(list)
    for l in db_learners:
        fn = (l.get("first_name") or "").strip()
        ln = (l.get("last_name") or "").strip()
        for variant in {norm_name(f"{ln} {fn}"), norm_name(f"{fn} {ln}")}:
            if variant:
                learner_by_name[variant].append(l)

    # 2. Fichiers source
    rows_sessions = read_xlsx(F_SESSIONS)
    rows_clients = read_xlsx(F_CLIENTS)
    rows_stag = read_xlsx(F_STAGIAIRES)

    # 2a. Code formation → entreprise (depuis fichier clients)
    ent_by_code = defaultdict(set)
    for r in rows_clients:
        code = norm(r.get("Code formation"))
        ent = norm(r.get("Entreprise/Client"))
        if code:
            ent_by_code[code].add(ent or "")

    # 2b. Code formation → session DB (depuis fichier sessions)
    code_to_session = {}          # code -> session DB (ou None)
    session_match = {"extid": 0, "title_date": 0, "title_only": 0, "none": 0}
    session_unmatched = []
    code_meta = {}                # code -> {title, start, heures, inter_intra}
    for r in rows_sessions:
        code = norm(r.get("Code formation"))
        if not code:
            continue
        title = norm(r.get("Nom de la formation"))
        start = to_date(r.get("Date de début de la formation"))
        code_meta[code] = {
            "title": title,
            "start": start,
            "heures_prevues": norm(r.get("Heures prévues")),
            "inter_intra": (norm(r.get("Inter/Intra/Autre")) or "").lower(),
        }
        extid = stable_external_id("session", title or "", start or "")
        s = sess_by_extid.get(extid)
        if s:
            session_match["extid"] += 1
        elif (norm_name(title), (start or "")) in sess_by_title_date:
            s = sess_by_title_date[(norm_name(title), start or "")]
            session_match["title_date"] += 1
        elif len(sess_by_title.get(norm_name(title), [])) == 1:
            s = sess_by_title[norm_name(title)][0]
            session_match["title_only"] += 1
        else:
            session_match["none"] += 1
            session_unmatched.append({"code": code, "title": title, "start": start,
                                      "homonymes_db": len(sess_by_title.get(norm_name(title), []))})
        code_to_session[code] = s

    # 2c. Code formation → client_id (entreprise → clients.company_name)
    code_to_client = {}           # code -> {"status","client_id"/"entreprise","n_entreprises"}
    client_match = {"intra_ok": 0, "intra_no_client": 0, "inter": 0, "no_entreprise": 0}
    client_unmatched_ent = defaultdict(int)
    for code in code_meta:
        ents = {e for e in ent_by_code.get(code, set()) if e}
        if not ents:
            code_to_client[code] = {"status": "no_entreprise"}
            client_match["no_entreprise"] += 1
            continue
        if len(ents) > 1:
            code_to_client[code] = {"status": "inter", "n_entreprises": len(ents)}
            client_match["inter"] += 1
            continue
        ent = next(iter(ents))
        cids = client_by_name.get(norm_name(ent), [])
        if len(cids) == 1:
            code_to_client[code] = {"status": "intra_ok", "client_id": cids[0], "entreprise": ent}
            client_match["intra_ok"] += 1
        else:
            code_to_client[code] = {"status": "intra_no_client", "entreprise": ent,
                                    "candidats": len(cids)}
            client_match["intra_no_client"] += 1
            client_unmatched_ent[ent] += 1

    # 2d. Inscriptions (stagaires) : Nom + Code formation → learner + session
    enr = {"total": 0, "session_ok_learner_ok": 0, "session_missing": 0,
           "learner_missing": 0, "learner_ambigu": 0}
    enr_learner_unmatched = []
    enr_client_resolvable = 0     # inscriptions dont le client_id est déductible (session INTRA matchée)
    for r in rows_stag:
        code = norm(r.get("Code formation"))
        nom = norm(r.get("Nom"))
        if not code or not nom:
            continue
        enr["total"] += 1
        s = code_to_session.get(code)
        cand = learner_by_name.get(norm_name(nom), [])
        if not s:
            enr["session_missing"] += 1
        if len(cand) == 0:
            enr["learner_missing"] += 1
            if len(enr_learner_unmatched) < 40:
                enr_learner_unmatched.append({"nom": nom, "code": code})
            continue
        if len(cand) > 1:
            enr["learner_ambigu"] += 1
            continue
        if s:
            enr["session_ok_learner_ok"] += 1
            cc = code_to_client.get(code, {})
            if cc.get("status") == "intra_ok":
                enr_client_resolvable += 1

    # ── Rapport ───────────────────────────────────────────────────────────
    report = {
        "generated_at": "(dry-run, lecture seule)",
        "db": {"sessions": len(db_sessions), "learners": len(db_learners), "clients": len(db_clients)},
        "codes_formation": len(code_meta),
        "jointure_session": session_match,
        "jointure_client": client_match,
        "inscriptions": enr,
        "inscriptions_client_id_resolvable": enr_client_resolvable,
        "entreprises_non_matchees": dict(sorted(client_unmatched_ent.items(),
                                                 key=lambda kv: -kv[1])[:30]),
        "sessions_non_matchees_sample": session_unmatched[:30],
        "learners_non_matches_sample": enr_learner_unmatched,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2))

    def pct(n, d):
        return f"{(100*n/d):.1f}%" if d else "—"

    print("═══ JOINTURE SESSION (Code formation → session DB) ═══")
    tot = len(code_meta)
    matched = session_match["extid"] + session_match["title_date"] + session_match["title_only"]
    print(f"  {len(rows_clients)} lignes clients | {tot} codes formation")
    print(f"  Matchées : {matched}/{tot} ({pct(matched, tot)})  "
          f"[extid={session_match['extid']}, titre+date={session_match['title_date']}, "
          f"titre seul={session_match['title_only']}]")
    print(f"  NON matchées : {session_match['none']}")

    print("\n═══ JOINTURE CLIENT (Code formation → client_id) ═══")
    print(f"  INTRA résolu (1 entreprise → 1 client) : {client_match['intra_ok']}")
    print(f"  INTRA entreprise sans client en base   : {client_match['intra_no_client']}")
    print(f"  INTER (>1 entreprise, par apprenant)   : {client_match['inter']}")
    print(f"  Sans entreprise dans la source         : {client_match['no_entreprise']}")

    print("\n═══ INSCRIPTIONS (stagaires : Nom + Code formation) ═══")
    print(f"  Total lignes               : {enr['total']}")
    print(f"  Session+learner OK         : {enr['session_ok_learner_ok']} ({pct(enr['session_ok_learner_ok'], enr['total'])})")
    print(f"  Apprenant introuvable      : {enr['learner_missing']}")
    print(f"  Apprenant homonyme (ambigu): {enr['learner_ambigu']}")
    print(f"  Session introuvable        : {enr['session_missing']}")
    print(f"  → client_id déductible (session INTRA matchée) : {enr_client_resolvable}")

    if report["entreprises_non_matchees"]:
        print("\n  Entreprises INTRA sans client en base (top) :")
        for e, n in report["entreprises_non_matchees"].items():
            print(f"    - {e} ({n})")
    if enr_learner_unmatched:
        print("\n  Apprenants non matchés (échantillon) :")
        for x in enr_learner_unmatched[:15]:
            print(f"    - {x['nom']} (code {x['code']})")

    print(f"\n✅ Rapport complet écrit dans {REPORT_PATH}")
    print("   (DRY-RUN : aucune écriture en base.)")


if __name__ == "__main__":
    main()
