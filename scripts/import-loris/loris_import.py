#!/usr/bin/env python3
"""
Import Loris XLSX → Supabase (MR FORMATION uniquement)

Voir scripts/import-loris/README.md pour l'usage complet.
"""

import argparse
import hashlib
import json
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

from openpyxl import load_workbook

# ── Config ────────────────────────────────────────────────────────────────

MR_ENTITY_ID = "f8acea54-71ab-4a22-8cf3-4e7170543bf1"
DOWNLOADS = Path.home() / "Downloads"
REPORT_PATH = Path(__file__).parent / "last_import_report.json"

FILES = {
    "clients": "Clients.xlsx",
    "learners": "Apprenants.xlsx",
    "trainings_sessions": "Suivi de l'activité.xlsx",
    "formation_trainers": "Suivi de l'activité des formateurs.xlsx",
    "enrollments": "Suivi de l'activité des stagaires.xlsx",
    "crm_quotes": "Suivi des devis.xlsx",
    "formation_invoices": "Suivi des factures.xlsx",
}

TABLE_ORDER = [
    "clients",
    "learners",
    "trainings_sessions",
    "formation_trainers",
    "enrollments",
    "crm_quotes",
    "formation_invoices",
]

# Charges LORIS (Type='Charge') — routées vers formation_charges, JAMAIS en factures.
# Préfixes PARTAGÉS avec scripts/import-loris/reclass_loris_charges.py (verrou d'ordre :
# tant que des factures « Loris Charge — » subsistent, l'import n'insère AUCUNE charge).
CHARGE_LABEL_PREFIX = "Charge LORIS — "   # labels formation_charges
CHARGE_NOTES_PREFIX = "Loris Charge — "   # notes des factures parasites historiques (pré-reclassement)

# Mapping conventionnel pour les colonnes Loris vs colonnes DB (gap → loris_metadata)

# ── Env loader ────────────────────────────────────────────────────────────

def load_env():
    env = {}
    env_path = Path(__file__).parent.parent.parent / ".env.local"
    if not env_path.exists():
        sys.exit(f"❌ .env.local introuvable à {env_path}")
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env


ENV = load_env()
SUPABASE_URL = ENV.get("NEXT_PUBLIC_SUPABASE_URL", "").rstrip("/")
SERVICE_ROLE = ENV.get("SUPABASE_SERVICE_ROLE_KEY", "")
if not SUPABASE_URL or not SERVICE_ROLE:
    sys.exit("❌ NEXT_PUBLIC_SUPABASE_URL ou SUPABASE_SERVICE_ROLE_KEY manquant")


# ── REST helpers ──────────────────────────────────────────────────────────

def _req(method, path, params=None, body=None, prefer=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    if params:
        url += "?" + urllib.parse.urlencode(params, quote_via=urllib.parse.quote)
    headers = {
        "apikey": SERVICE_ROLE,
        "Authorization": f"Bearer {SERVICE_ROLE}",
    }
    if body is not None:
        headers["Content-Type"] = "application/json"
    if prefer:
        headers["Prefer"] = prefer
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as r:
            raw = r.read()
            if not raw:
                return None
            return json.loads(raw)
    except urllib.error.HTTPError as e:
        body_txt = e.read().decode(errors="replace")
        return {"_error": True, "status": e.code, "body": body_txt}


def rest_get(table, **params):
    return _req("GET", table, params=params)


def rest_post(table, rows, prefer="return=representation"):
    return _req("POST", table, body=rows, prefer=prefer)


def rest_get_all(table, soft=False, **params):
    """GET paginé — PostgREST plafonne ~1000 lignes/réponse.

    Avance de len(page) et s'arrête sur page VIDE (robuste même si le serveur
    plafonne sous la limite demandée). Sur _error : sys.exit par défaut (pas
    d'échec silencieux — un GET de dédupe/verrou qui échoue en silence = doublons
    garantis) ; soft=True retourne None à la place, pour les appels situés APRÈS
    des écritures (le rapport JSON du run doit toujours être écrit)."""
    rows = []
    offset = 0
    base = dict(params)
    base.setdefault("order", "id.asc")
    base.setdefault("limit", "1000")
    while True:
        page = _req("GET", table, params={**base, "offset": str(offset)})
        if isinstance(page, dict) and page.get("_error"):
            if soft:
                print(f"  ⚠️  Erreur GET {table} : HTTP {page['status']} — {page['body'][:200]}")
                return None
            sys.exit(f"❌ Erreur GET {table} : HTTP {page['status']} — {page['body'][:300]}")
        page = page if isinstance(page, list) else []
        rows.extend(page)
        if not page:
            return rows
        offset += len(page)


# ── XLSX → dicts ──────────────────────────────────────────────────────────

def read_xlsx(filename):
    path = DOWNLOADS / filename
    if not path.exists():
        return None, []
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, []
    headers = [(str(h).strip() if h is not None else f"col_{i}") for i, h in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        d = {}
        for i, val in enumerate(row):
            if i < len(headers):
                d[headers[i]] = val
        if any(v is not None and str(v).strip() for v in d.values()):
            data.append(d)
    wb.close()
    return headers, data


# ── Helpers ───────────────────────────────────────────────────────────────

def stable_external_id(prefix, *parts):
    """Hash stable des champs clés pour idempotence (16 hex chars)."""
    h = hashlib.sha256("|".join(str(p or "") for p in parts).encode()).hexdigest()[:16]
    return f"loris-{prefix}-{h}"


def norm(v):
    """Normalise une valeur : strip + None si vide ou tiret."""
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in ("-", "—", "N/A", "NA"):
        return None
    return s


def to_date(v):
    """Convertit en YYYY-MM-DD string."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    # Already YYYY-MM-DD
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    # DD/MM/YYYY
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    return None


def to_decimal(v):
    """Convertit en float (gère '1,200.00 EUR', '500', etc.)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = re.sub(r"[^\d.,\-]", "", s)
    if not s:
        return None
    # If both . and ,, last separator decides decimal
    if "." in s and "," in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def norm_name(s):
    """Normalise un nom pour matching : lowercase, espaces multiples → 1, NFD strip accents."""
    if s is None:
        return ""
    import unicodedata
    s = str(s).strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s


def split_name(full):
    """Sépare 'NOM Prénom' (NOM en majuscules) → (first, last). Tolère noms 1-mot
    en mettant le nom complet dans last_name et un placeholder dans first_name
    (la colonne DB exige first_name NOT NULL)."""
    if not full:
        return None, None
    full = full.strip()
    parts = full.split()
    if len(parts) == 1:
        # Un seul token (ex: "J.DENIS", "MARTINEAU") — préserver dans last_name
        return "—", parts[0]
    # If first token is all-uppercase, treat as last name
    if parts and parts[0].isupper() and len(parts[0]) > 1:
        last = parts[0]
        first = " ".join(parts[1:])
    else:
        first = parts[0]
        last = " ".join(parts[1:])
    return first, last


# ── MAPPERS ───────────────────────────────────────────────────────────────

def map_client(row, idx):
    name = norm(row.get("Nom")) or f"Client Loris {idx}"
    email = norm(row.get("Email"))
    metadata_keys = [
        "Représenté par", "Fonction du représentant", "TVA", "APE",
        "OPCO", "Numéro d'adhérent à l'OPCO", "Effective de l'entreprise",
        "IDCC", "NACE",
        "Contact 2", "Email 2", "Contact 3", "Email 3",
        "Contact 4", "Email 4", "Contact 5", "Email 5",
    ]
    metadata = {k: row[k] for k in metadata_keys if norm(row.get(k))}
    ext_id = norm(row.get("ID Externe")) or stable_external_id("client", name, email or "")
    return {
        "entity_id": MR_ENTITY_ID,
        "company_name": name,
        "phone": norm(row.get("Tel")),
        "email": email,
        "address": norm(row.get("Adresse")),
        "siret": norm(row.get("SIRET")),
        "postal_code": norm(row.get("Code postal")),
        "notes": norm(row.get("Description")),
        "opco": norm(row.get("OPCO")),
        "loris_external_id": ext_id,
        "loris_metadata": metadata,
    }


def map_learner(row, idx, clients_by_name):
    nom_field = norm(row.get("Nom"))
    first, last = split_name(nom_field)
    if not first and not last:
        return None
    email = norm(row.get("Email"))
    entreprise = norm(row.get("Entreprise"))
    client_id = clients_by_name.get(norm_name(entreprise)) if entreprise else None

    # birth fields
    birth_date = to_date(row.get("Date de naissance"))
    birth_city = norm(row.get("Ville de naissance"))

    # sex mapping
    sex_raw = norm(row.get("Sexe"))
    gender = None
    if sex_raw:
        sx = sex_raw.lower()
        if sx.startswith("h") or sx.startswith("m"):
            gender = "M"
        elif sx.startswith("f"):
            gender = "F"

    metadata_keys = [
        "Sessions", "Département de naissance", "Pays de naissance",
        "Description", "Statut", "Fonction", "Profession", "Profession 2",
        "Raison Sociale", "Reconnaissance Travailleur Handicapé",
        "Catégorie socio-professionnelle", "Nature du contrat de travail",
        "Salaire Horaire Brut", "Date de création",
    ]
    metadata = {k: row[k] for k in metadata_keys if norm(row.get(k))}
    if entreprise and not client_id:
        metadata["_unmatched_entreprise"] = entreprise

    ext_id = norm(row.get("ID Externe")) or stable_external_id(
        "learner", first or "", last or "", email or ""
    )

    return {
        "entity_id": MR_ENTITY_ID,
        "client_id": client_id,
        "first_name": first,
        "last_name": last,
        "email": email,
        "phone": norm(row.get("Tel")),
        "job_title": norm(row.get("Fonction")),
        "address": norm(row.get("Adresse")),
        "birth_date": birth_date,
        "birth_city": birth_city,
        "gender": gender,
        "social_security_number": norm(row.get("No. Sécurité Sociale")),
        "loris_external_id": ext_id,
        "loris_metadata": metadata,
    }


def map_training_from_session(row):
    """Extrait une training (formation au sens catalogue) depuis une ligne 'Suivi de l'activité'."""
    title = norm(row.get("Nom de la formation"))
    if not title:
        return None
    duration_h = None
    hours_raw = row.get("Heures prévues")
    if isinstance(hours_raw, str) and ":" in hours_raw:
        parts = hours_raw.split(":")
        duration_h = int(parts[0]) if parts[0].isdigit() else None
    metadata = {
        "loris_inter_intra": norm(row.get("Inter/Intra/Autre")),
    }
    metadata = {k: v for k, v in metadata.items() if v}
    return {
        "entity_id": MR_ENTITY_ID,
        "title": title,
        "duration_hours": duration_h,
        "is_active": True,
        "loris_external_id": stable_external_id("training", title),
        "loris_metadata": metadata,
    }


def map_session(row, idx, trainings_by_title):
    title = norm(row.get("Nom de la formation"))
    if not title:
        return None
    training_id = trainings_by_title.get(norm_name(title))

    start_date = to_date(row.get("Date de début de la formation"))
    end_date = to_date(row.get("Date de fin de la formation"))
    amount_ht = to_decimal(row.get("Montant HT"))
    charges = to_decimal(row.get("Charges HT"))
    location = norm(row.get("Emplacement"))

    status_raw = (norm(row.get("Statut")) or "").lower()
    if "planif" in status_raw:
        status_db = "planned"
    elif "termin" in status_raw or "achev" in status_raw:
        status_db = "completed"
    elif "cours" in status_raw:
        status_db = "in_progress"
    else:
        status_db = "planned"

    metadata = {
        "loris_apprenants_text": norm(row.get("Apprenants")),
        "loris_entreprises_text": norm(row.get("Entreprises")),
        "loris_formateurs_text": norm(row.get("Formateurs")),
        "loris_financeurs": norm(row.get("Financeurs")),
        "loris_manager": norm(row.get("Manager")),
        "loris_charges_ht": charges,
        "loris_facture": norm(row.get("Facturé")),
        "loris_inter_intra": norm(row.get("Inter/Intra/Autre")),
        "loris_date_achevement": to_date(row.get("Date d'achèvement de la formation")),
    }
    metadata = {k: v for k, v in metadata.items() if v is not None}

    return {
        "entity_id": MR_ENTITY_ID,
        "training_id": training_id,
        "title": title,
        "start_date": start_date,
        "end_date": end_date,
        "status": status_db,
        "location": location,
        "price": amount_ht,
        "total_price": amount_ht,
        "loris_external_id": stable_external_id("session", title, start_date or ""),
        "loris_metadata": metadata,
    }


def map_formation_trainer(row, sessions_by_title, trainers_by_name, code_to_session=None):
    """Matching session par Code formation (désambiguïse les titres identiques), repli sur le titre."""
    formateur_name = norm(row.get("Formateur"))
    title = norm(row.get("Formation"))
    if not formateur_name or not title:
        return None
    start_date = to_date(row.get("Date de début"))
    code = norm(row.get("Code formation"))
    session_id = code_to_session.get(code) if (code and code_to_session) else None
    if not session_id:
        session_id = sessions_by_title.get(norm_name(title))
    trainer_id = trainers_by_name.get(norm_name(formateur_name))
    if not session_id or not trainer_id:
        return {
            "_skip_reason": f"no_session_or_trainer (session={session_id}, trainer={trainer_id})",
            "_meta": {"formateur": formateur_name, "title": title, "start_date": start_date},
        }
    # Parse hours "07h00" → 7
    def parse_hours(s):
        if not s:
            return None
        m = re.match(r"^(\d+)h(\d+)?", str(s))
        if m:
            return int(m.group(1)) + (int(m.group(2)) / 60 if m.group(2) else 0)
        return None
    tarif = to_decimal(row.get("Tarif"))
    par = norm(row.get("Par")) or ""
    is_daily = "jour" in par.lower()

    # PostgREST batch insert exige des keys identiques — toujours fournir les 2 colonnes
    payload = {
        "session_id": session_id,
        "trainer_id": trainer_id,
        "role": "formateur",
        "hours_done": parse_hours(row.get("Heures réalisées")),
        "hourly_rate": tarif if (tarif is not None and not is_daily) else None,
        "daily_rate": tarif if (tarif is not None and is_daily) else None,
        "loris_external_id": stable_external_id("ft", title, formateur_name, start_date or ""),
        "loris_metadata": {
            "loris_par": par,
            "loris_heures_prevues": norm(row.get("Heures prévues")),
        },
    }
    return payload


def map_enrollment(row, sessions_by_title, learners_by_name, code_to_session=None):
    formation = norm(row.get("Formation"))
    nom = norm(row.get("Nom"))
    if not formation or not nom:
        return None
    # Matching session : par Code formation (fiable, désambiguïse les titres identiques),
    # repli sur le titre si le code est absent/non résolu. (Fix bug d'inscriptions mal attribuées.)
    code = norm(row.get("Code formation"))
    session_id = code_to_session.get(code) if (code and code_to_session) else None
    if not session_id:
        session_id = sessions_by_title.get(norm_name(formation))
    learner_id = learners_by_name.get(norm_name(nom))
    if not session_id or not learner_id:
        return {
            "_skip_reason": f"no_session_or_learner (session={session_id}, learner={learner_id})",
            "_meta": {"formation": formation, "nom": nom},
        }
    completion_raw = norm(row.get("Heures réalisées"))
    # CHECK constraint enrollments.status : ('registered','confirmed','cancelled','completed')
    status_db = "completed" if completion_raw and completion_raw != "00:00:00" else "registered"
    abandon = norm(row.get("Abandon/Absences non justifiées sans reprise"))
    if abandon:
        status_db = "cancelled"

    metadata = {
        "loris_heures_prevues": norm(row.get("Heures prévues")),
        "loris_heures_realisees": completion_raw,
        "loris_elearning_start": to_date(row.get("Date de début de l'e-learning")),
        "loris_elearning_end": to_date(row.get("Date de fin de l'e-learning")),
        "loris_heures_elearning": norm(row.get("Heures réalisées en e-learning")),
        "loris_notes": norm(row.get("Notes")),
        "loris_formateurs": norm(row.get("Formateurs")),
        "loris_date_creation": to_date(row.get("Date de Création")),
        "loris_abandon": abandon,
    }
    metadata = {k: v for k, v in metadata.items() if v is not None}

    price = to_decimal(row.get("Prix total"))
    payload = {
        "session_id": session_id,
        "learner_id": learner_id,
        "status": status_db,
        "individual_price": price,
        "loris_external_id": stable_external_id("enr", formation, nom),
        "loris_metadata": metadata,
    }
    return payload


def map_crm_quote(row, idx, clients_by_name):
    reference = norm(row.get("Numéro")) or f"LORIS-DEVIS-{idx}"
    destinataire = norm(row.get("Destinataire"))
    client_id = clients_by_name.get(norm_name(destinataire)) if destinataire else None

    # CHECK constraint crm_quotes.status : ('draft','sent','accepted','rejected','expired')
    status_raw = (norm(row.get("Statut")) or "").lower()
    if "accept" in status_raw:
        status_db = "accepted"
    elif "retard" in status_raw or "expir" in status_raw:
        status_db = "expired"
    elif "envoy" in status_raw:
        status_db = "sent"
    elif "refus" in status_raw:
        status_db = "rejected"
    else:
        status_db = "draft"

    amount = to_decimal(row.get("Montant"))
    date_created = to_date(row.get("Date"))
    valid_until = to_date(row.get("Date d'échéance"))

    metadata = {
        "loris_destinataire": destinataire,
        "loris_nom_prospect": norm(row.get("Nom du prospect")),
        "loris_type": norm(row.get("Type")),
        "loris_status_raw": norm(row.get("Statut")),
    }
    metadata = {k: v for k, v in metadata.items() if v is not None}

    # quote_number est INTEGER en DB — on utilise idx (la reference texte est dans `reference`)
    return {
        "entity_id": MR_ENTITY_ID,
        "reference": reference,
        "quote_number": 900000 + idx,  # plage haute pour ne pas collisionner les vrais devis
        "client_id": client_id,
        "amount": amount,
        "status": status_db,
        "valid_until": valid_until,
        "created_at": date_created + "T00:00:00Z" if date_created else None,
        "loris_external_id": stable_external_id("quote", reference),
        "loris_metadata": metadata,
    }


def map_formation_invoice(row, idx, sessions_by_title, clients_by_name, code_to_session=None):
    reference = norm(row.get("Numéro")) or f"LORIS-FAC-{idx}"
    destinataire = norm(row.get("Destinataire")) or norm(row.get("Client"))
    formation_title = norm(row.get("Nom de la formation"))

    # Matching session : par Code formation (fiable, désambiguïse les titres identiques),
    # repli sur le titre si le code est absent/non résolu. (Fix factures mal attribuées.)
    code = norm(row.get("Code formation"))
    session_id = code_to_session.get(code) if (code and code_to_session) else None
    if not session_id and formation_title:
        session_id = sessions_by_title.get(norm_name(formation_title))
    if not session_id:
        return {
            "_skip_reason": f"no_session for formation '{formation_title}'",
            "_meta": {"reference": reference, "destinataire": destinataire},
        }

    amount = to_decimal(row.get("Montant"))
    status_raw = (norm(row.get("Statut")) or "").lower()
    if "pay" in status_raw:
        status_db = "paid"
    elif "retard" in status_raw:
        status_db = "late"
    elif "envoy" in status_raw or "sent" in status_raw:
        status_db = "sent"
    elif "annul" in status_raw:
        status_db = "cancelled"
    else:
        status_db = "pending"

    # Recipient mapping : guess if it's a learner (no entreprise match) or company
    client_id = clients_by_name.get(norm_name(destinataire)) if destinataire else None
    recipient_type = "company" if client_id else "learner"
    recipient_id = client_id  # if recipient_type=learner and no id, this will fail — fallback to placeholder UUID
    recipient_name = destinataire or "Inconnu"

    if recipient_id is None:
        return {
            "_skip_reason": f"no_recipient_id for '{destinataire}' (type={recipient_type})",
            "_meta": {"reference": reference, "destinataire": destinataire, "amount": amount},
        }

    # number + global_number sont INTEGER NOT NULL — plages hautes pour ne pas collisionner
    payload = {
        "entity_id": MR_ENTITY_ID,
        "session_id": session_id,
        "recipient_type": recipient_type,
        "recipient_id": recipient_id,
        "recipient_name": recipient_name,
        "amount": amount or 0,
        "prefix": "LORIS",
        "number": 900000 + idx,        # unique per import — plage haute
        "global_number": 900000 + idx,  # NOT NULL ajouté par migration add_invoice_global_numbering
        "status": status_db,
        "due_date": to_date(row.get("Date d'échéance")),
        "paid_at": (to_date(row.get("Date de paiement")) + "T00:00:00Z") if to_date(row.get("Date de paiement")) else None,
        "external_reference": reference,
        "external_source": "loris",
        "is_external": True,
        "notes": f"Loris {norm(row.get('Type')) or 'Facture'} — mode={norm(row.get('Mode de paiement')) or 'n/a'}",
    }
    return payload


# ── Charges (Type='Charge') → formation_charges ──────────────────────────

def canon_charge_amount(v):
    """Montant canonique de dédupe : valeur absolue quantizée half-up à 2 décimales
    (None si NULL/illisible/NaN).

    Decimal + ROUND_HALF_UP obligatoire : le round() banker's de Python divergerait
    du NUMERIC Postgres sur les demi-cents. Guard NaN AVANT quantize (d != d ⇔ NaN) :
    Decimal('NaN').quantize lèverait InvalidOperation et NaN casserait les clés de
    dédupe (jamais égal à lui-même)."""
    d = to_decimal(v)
    if d is None or d != d or d in (float("inf"), float("-inf")):  # NULL, NaN (d != d), ±inf
        return None
    return abs(Decimal(str(d))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def charge_label_base(label):
    """Base de dédupe d'une charge EXISTANTE en base : label sans le préfixe
    'Charge LORIS — ' ni le suffixe ' ({référence})' FINAL, normalisé via
    norm_name (insensible casse/accents).

    Regex ancrée sur le suffixe parenthésé FINAL — PAS split(' (', 1), qui
    tronquerait un destinataire contenant lui-même ' ('."""
    s = label or ""
    if s.startswith(CHARGE_LABEL_PREFIX):
        s = s[len(CHARGE_LABEL_PREFIX):]
    s = re.sub(r"\s*\([^()]*\)$", "", s)
    return norm_name(s)


def map_formation_charge(row, sessions_by_title, code_to_session=None):
    """Ligne xlsx Type='Charge' (coût formateur/fournisseur) → payload formation_charges.

    Label : 'Charge LORIS — {Destinataire} ({Numéro})' si Numéro présent, sinon
    'Charge LORIS — {Destinataire}' — JAMAIS de référence synthétique dépendante
    de l'index (non stable entre runs). L'idempotence NE repose PAS sur le label :
    dédupe MULTISET par (entity_id, session_id, norm_name(Destinataire), montant
    quantizé) dans la sous-étape charges de main(). Le payload embarque
    '_dedupe_base' (retiré avant insertion)."""
    destinataire = norm(row.get("Destinataire")) or norm(row.get("Client"))
    formation_title = norm(row.get("Nom de la formation"))

    # Matching session : identique à map_formation_invoice (Code formation fiable,
    # repli sur le titre si le code est absent/non résolu).
    code = norm(row.get("Code formation"))
    session_id = code_to_session.get(code) if (code and code_to_session) else None
    if not session_id and formation_title:
        session_id = sessions_by_title.get(norm_name(formation_title))
    if not session_id:
        return {
            "_skip_reason": f"no_session for formation '{formation_title}'",
            "_meta": {"destinataire": destinataire},
        }

    amount = to_decimal(row.get("Montant"))
    if (amount is None or amount != amount            # None ou NaN (d != d)
            or amount in (float("inf"), float("-inf"))
            or abs(amount) >= 10**8):                 # déborderait DECIMAL(10,2) → rejet du batch entier
        return {
            "_skip_reason": f"montant illisible/hors bornes ('{row.get('Montant')}') pour charge '{destinataire}'",
            "_meta": {"destinataire": destinataire},
        }

    reference = norm(row.get("Numéro"))
    recipient = destinataire or "Inconnu"
    if reference:
        label = f"{CHARGE_LABEL_PREFIX}{recipient} ({reference})"
    else:
        label = f"{CHARGE_LABEL_PREFIX}{recipient}"

    # formation_charges : id, session_id, entity_id, label, amount, created_at — rien d'autre.
    # _dedupe_base : calculée depuis le label QUI SERA ÉCRIT, via charge_label_base —
    # la même fonction que pour les charges existantes → symétrie par construction
    # (un destinataire finissant par « (...) » donne la même base des deux côtés).
    return {
        "session_id": session_id,
        "entity_id": MR_ENTITY_ID,
        "label": label,
        "amount": abs(amount),
        "_dedupe_base": charge_label_base(label),
    }


# ── Import orchestrator ───────────────────────────────────────────────────

def fetch_existing_lookups():
    """Récupère les mappings name → id depuis Supabase pour matching."""
    out = {}

    print("  📋 Fetch clients existants...")
    rows = rest_get("clients", select="id,company_name,email,loris_external_id",
                    **{"entity_id": f"eq.{MR_ENTITY_ID}"})
    rows = rows if isinstance(rows, list) else []
    out["clients_by_name"] = {norm_name(r["company_name"]): r["id"] for r in rows if r.get("company_name")}
    out["clients_by_email"] = {norm_name(r["email"]): r["id"] for r in rows if r.get("email")}
    out["clients_by_loris_id"] = {r["loris_external_id"]: r["id"] for r in rows if r.get("loris_external_id")}

    print("  📋 Fetch learners existants...")
    rows = rest_get("learners", select="id,first_name,last_name,email,loris_external_id",
                    **{"entity_id": f"eq.{MR_ENTITY_ID}"})
    rows = rows if isinstance(rows, list) else []
    out["learners_by_email"] = {norm_name(r["email"]): r["id"] for r in rows if r.get("email")}
    out["learners_by_loris_id"] = {r["loris_external_id"]: r["id"] for r in rows if r.get("loris_external_id")}
    out["learners_by_name"] = {}
    for r in rows:
        fn = r.get("first_name") or ""
        ln = r.get("last_name") or ""
        # Multiple variants for fuzzy matching
        for variant in [f"{ln} {fn}", f"{fn} {ln}", ln, fn]:
            k = norm_name(variant)
            if k:
                out["learners_by_name"][k] = r["id"]

    print("  📋 Fetch trainers existants...")
    rows = rest_get("trainers", select="id,first_name,last_name",
                    **{"entity_id": f"eq.{MR_ENTITY_ID}"})
    rows = rows if isinstance(rows, list) else []
    out["trainers_by_name"] = {}
    for r in rows:
        fn = r.get("first_name") or ""
        ln = r.get("last_name") or ""
        for variant in [f"{ln} {fn}", f"{fn} {ln}", ln, fn]:
            k = norm_name(variant)
            if k:
                out["trainers_by_name"][k] = r["id"]

    print("  📋 Fetch trainings existants...")
    rows = rest_get("trainings", select="id,title,loris_external_id",
                    **{"entity_id": f"eq.{MR_ENTITY_ID}"})
    rows = rows if isinstance(rows, list) else []
    out["trainings_by_title"] = {norm_name(r["title"]): r["id"] for r in rows if r.get("title")}
    out["trainings_by_loris_id"] = {r["loris_external_id"]: r["id"] for r in rows if r.get("loris_external_id")}

    print("  📋 Fetch sessions existantes...")
    rows = rest_get("sessions", select="id,title,start_date,loris_external_id",
                    **{"entity_id": f"eq.{MR_ENTITY_ID}"})
    rows = rows if isinstance(rows, list) else []
    # sessions_by_title : first session matching (assouplit le matching par date)
    out["sessions_by_title"] = {}
    out["sessions_by_title_date"] = {}
    for r in rows:
        if r.get("title"):
            k = norm_name(r["title"])
            out["sessions_by_title"].setdefault(k, r["id"])
            out["sessions_by_title_date"][(k, r.get("start_date"))] = r["id"]
    out["sessions_by_loris_id"] = {r["loris_external_id"]: r["id"] for r in rows if r.get("loris_external_id")}

    # code_to_session : Code formation (fichier activité) → session_id, via extid (titre+date début).
    # Construit depuis les sessions DÉJÀ en base ; complété au fil des insertions de sessions (cf. boucle).
    out["code_to_session"] = {}
    _sh, _sdata = read_xlsx(FILES["trainings_sessions"])
    for _row in (_sdata or []):
        _code = norm(_row.get("Code formation"))
        _title = norm(_row.get("Nom de la formation"))
        _start = to_date(_row.get("Date de début de la formation"))
        if not _code or not _title:
            continue
        _sid = out["sessions_by_loris_id"].get(stable_external_id("session", _title, _start or ""))
        if _sid:
            out["code_to_session"][_code] = _sid

    return out


def insert_batch(table, rows, dry_run, batch_size=100, ignore_duplicates=False):
    """Insert with auto-batching. Returns (inserted_ids, errors)."""
    if dry_run:
        return [], []
    inserted = []
    errors = []
    prefer = "return=representation"
    if ignore_duplicates:
        prefer += ",resolution=ignore-duplicates"
    for i in range(0, len(rows), batch_size):
        batch = rows[i : i + batch_size]
        result = rest_post(table, batch, prefer=prefer)
        if isinstance(result, dict) and result.get("_error"):
            errors.append({"batch_start": i, "status": result["status"], "body": result["body"][:500]})
        elif isinstance(result, list):
            inserted.extend(result)
    return inserted, errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Analyse seule, n'écrit rien")
    parser.add_argument("--execute", action="store_true", help="Exécute pour de vrai")
    parser.add_argument("--tables", default="all", help="liste comma-séparée ou 'all'")
    args = parser.parse_args()

    if not args.dry_run and not args.execute:
        sys.exit("❌ Précise --dry-run ou --execute")
    if args.dry_run and args.execute:
        sys.exit("❌ --dry-run et --execute mutuellement exclusifs")

    dry = args.dry_run
    selected = TABLE_ORDER if args.tables == "all" else args.tables.split(",")

    mode = "DRY-RUN" if dry else "EXECUTE"
    print(f"\n{'═'*70}\n🚀 Loris Import — Mode {mode}\n{'═'*70}")
    print(f"Tables ciblées : {selected}")
    print(f"Entity MR FORMATION : {MR_ENTITY_ID}\n")

    # 1. Lookups initiaux
    print("📥 Récupération des lookups existants...")
    lk = fetch_existing_lookups()
    print(f"  → {len(lk['clients_by_name'])} clients, {len(lk['learners_by_name'])} learners, "
          f"{len(lk['trainers_by_name'])} trainers, {len(lk['trainings_by_title'])} trainings, "
          f"{len(lk['sessions_by_title'])} sessions\n")

    report = {"mode": mode, "started_at": datetime.utcnow().isoformat() + "Z", "tables": {}}

    # 2. CLIENTS
    if "clients" in selected:
        print(f"\n{'─'*70}\n📂 CLIENTS\n{'─'*70}")
        headers, data = read_xlsx(FILES["clients"])
        if not data:
            print(f"  ⚠️  {FILES['clients']} introuvable ou vide")
        else:
            to_insert, skipped_dup = [], 0
            seen_ext = set()
            for idx, row in enumerate(data):
                payload = map_client(row, idx)
                if not payload.get("company_name"):
                    continue
                # Dedup : existing in DB OR already seen in batch
                if payload["loris_external_id"] in lk["clients_by_loris_id"]:
                    skipped_dup += 1
                    continue
                if payload["email"] and norm_name(payload["email"]) in lk["clients_by_email"]:
                    skipped_dup += 1
                    continue
                if norm_name(payload["company_name"]) in lk["clients_by_name"]:
                    skipped_dup += 1
                    continue
                if payload["loris_external_id"] in seen_ext:
                    skipped_dup += 1
                    continue
                seen_ext.add(payload["loris_external_id"])
                to_insert.append(payload)

            print(f"  → {len(to_insert)} à insérer, {skipped_dup} skippés (déjà existants)")
            inserted, errors = insert_batch("clients", to_insert, dry)
            print(f"  ✅ inserted={len(inserted)} | errors={len(errors)}")
            if errors:
                print(f"  Sample error : {errors[0]}")
            # Refresh lookups
            for r in inserted:
                if r.get("company_name"):
                    lk["clients_by_name"][norm_name(r["company_name"])] = r["id"]
                if r.get("loris_external_id"):
                    lk["clients_by_loris_id"][r["loris_external_id"]] = r["id"]
            report["tables"]["clients"] = {
                "total_rows": len(data),
                "to_insert": len(to_insert),
                "skipped_duplicates": skipped_dup,
                "inserted": len(inserted),
                "errors": len(errors),
                "error_samples": errors[:3],
            }

    # 3. LEARNERS
    if "learners" in selected:
        print(f"\n{'─'*70}\n📂 LEARNERS\n{'─'*70}")
        headers, data = read_xlsx(FILES["learners"])
        if not data:
            print(f"  ⚠️  {FILES['learners']} introuvable")
        else:
            to_insert, skipped_dup, skipped_invalid = [], 0, 0
            seen_ext = set()
            for idx, row in enumerate(data):
                payload = map_learner(row, idx, lk["clients_by_name"])
                if not payload:
                    skipped_invalid += 1
                    continue
                if payload["loris_external_id"] in lk["learners_by_loris_id"]:
                    skipped_dup += 1
                    continue
                if payload["email"] and norm_name(payload["email"]) in lk["learners_by_email"]:
                    skipped_dup += 1
                    continue
                name_key = norm_name(f"{payload.get('last_name') or ''} {payload.get('first_name') or ''}")
                if name_key and name_key in lk["learners_by_name"]:
                    skipped_dup += 1
                    continue
                if payload["loris_external_id"] in seen_ext:
                    skipped_dup += 1
                    continue
                seen_ext.add(payload["loris_external_id"])
                to_insert.append(payload)

            print(f"  → {len(to_insert)} à insérer, {skipped_dup} skippés doublons, {skipped_invalid} invalides")
            inserted, errors = insert_batch("learners", to_insert, dry)
            print(f"  ✅ inserted={len(inserted)} | errors={len(errors)}")
            if errors:
                print(f"  Sample error : {errors[0]}")
            for r in inserted:
                name_key = norm_name(f"{r.get('last_name') or ''} {r.get('first_name') or ''}")
                if name_key:
                    lk["learners_by_name"][name_key] = r["id"]
                if r.get("loris_external_id"):
                    lk["learners_by_loris_id"][r["loris_external_id"]] = r["id"]
            report["tables"]["learners"] = {
                "total_rows": len(data),
                "to_insert": len(to_insert),
                "skipped_duplicates": skipped_dup,
                "skipped_invalid": skipped_invalid,
                "inserted": len(inserted),
                "errors": len(errors),
                "error_samples": errors[:3],
            }

    # 4. TRAININGS + SESSIONS (depuis Suivi activité)
    if "trainings_sessions" in selected:
        print(f"\n{'─'*70}\n📂 TRAININGS + SESSIONS\n{'─'*70}")
        headers, data = read_xlsx(FILES["trainings_sessions"])
        if not data:
            print(f"  ⚠️  {FILES['trainings_sessions']} introuvable")
        else:
            # Trainings : unique par title
            trainings_dedup = {}
            for row in data:
                t = map_training_from_session(row)
                if t and norm_name(t["title"]) not in trainings_dedup:
                    trainings_dedup[norm_name(t["title"])] = t

            to_insert_t = []
            skipped_dup_t = 0
            for t in trainings_dedup.values():
                if norm_name(t["title"]) in lk["trainings_by_title"]:
                    skipped_dup_t += 1
                    continue
                to_insert_t.append(t)
            print(f"  Trainings → {len(to_insert_t)} à insérer ({skipped_dup_t} doublons)")
            inserted_t, errors_t = insert_batch("trainings", to_insert_t, dry)
            print(f"  ✅ trainings inserted={len(inserted_t)} | errors={len(errors_t)}")
            if errors_t:
                print(f"  Sample : {errors_t[0]}")
            for r in inserted_t:
                if r.get("title"):
                    lk["trainings_by_title"][norm_name(r["title"])] = r["id"]

            # Sessions : 1 par ligne du fichier (peut avoir plusieurs sessions pour la même training)
            to_insert_s, skipped_dup_s = [], 0
            seen_ext = set()
            for idx, row in enumerate(data):
                s = map_session(row, idx, lk["trainings_by_title"])
                if not s:
                    continue
                if s["loris_external_id"] in lk["sessions_by_loris_id"]:
                    skipped_dup_s += 1
                    continue
                key = (norm_name(s["title"]), s.get("start_date"))
                if key in lk["sessions_by_title_date"]:
                    skipped_dup_s += 1
                    continue
                if s["loris_external_id"] in seen_ext:
                    skipped_dup_s += 1
                    continue
                seen_ext.add(s["loris_external_id"])
                to_insert_s.append(s)
            print(f"  Sessions → {len(to_insert_s)} à insérer ({skipped_dup_s} doublons)")
            inserted_s, errors_s = insert_batch("sessions", to_insert_s, dry)
            print(f"  ✅ sessions inserted={len(inserted_s)} | errors={len(errors_s)}")
            if errors_s:
                print(f"  Sample : {errors_s[0]}")
            for r in inserted_s:
                if r.get("title"):
                    lk["sessions_by_title"][norm_name(r["title"])] = r["id"]
                    lk["sessions_by_title_date"][(norm_name(r["title"]), r.get("start_date"))] = r["id"]
                if r.get("loris_external_id"):
                    lk["sessions_by_loris_id"][r["loris_external_id"]] = r["id"]
            # Compléter code_to_session avec les sessions du fichier (fraîchement insérées ou existantes).
            for row in data:
                _code = norm(row.get("Code formation"))
                _title = norm(row.get("Nom de la formation"))
                _start = to_date(row.get("Date de début de la formation"))
                if not _code or not _title:
                    continue
                _sid = lk["sessions_by_loris_id"].get(stable_external_id("session", _title, _start or ""))
                if _sid:
                    lk["code_to_session"][_code] = _sid

            report["tables"]["trainings_sessions"] = {
                "total_rows": len(data),
                "trainings_inserted": len(inserted_t),
                "trainings_skipped": skipped_dup_t,
                "trainings_errors": len(errors_t),
                "sessions_inserted": len(inserted_s),
                "sessions_skipped": skipped_dup_s,
                "sessions_errors": len(errors_s),
                "error_samples": (errors_t + errors_s)[:3],
            }

    # 5. FORMATION_TRAINERS
    if "formation_trainers" in selected:
        print(f"\n{'─'*70}\n📂 FORMATION_TRAINERS\n{'─'*70}")
        headers, data = read_xlsx(FILES["formation_trainers"])
        if not data:
            print(f"  ⚠️  {FILES['formation_trainers']} introuvable")
        else:
            # Bootstrap : créer les trainers Loris manquants en bulk avant le mapping
            loris_trainers = set()
            for row in data:
                fn = norm(row.get("Formateur"))
                if fn:
                    loris_trainers.add(fn)
            missing = [t for t in loris_trainers if norm_name(t) not in lk["trainers_by_name"]]
            if missing:
                print(f"  🔧 Bootstrap : {len(missing)} trainers Loris à créer en pré-import")
                trainer_payloads = []
                for name in missing:
                    first, last = split_name(name)
                    trainer_payloads.append({
                        "entity_id": MR_ENTITY_ID,
                        "first_name": first or "—",
                        "last_name": last or name,
                        "type": "external",
                    })
                if not dry:
                    inserted_trainers, errs = insert_batch("trainers", trainer_payloads, dry)
                    for r in inserted_trainers:
                        fn = r.get("first_name") or ""
                        ln = r.get("last_name") or ""
                        for variant in [f"{ln} {fn}", f"{fn} {ln}", ln, fn]:
                            k = norm_name(variant)
                            if k:
                                lk["trainers_by_name"][k] = r["id"]
                    print(f"  ✅ Bootstrap trainers : {len(inserted_trainers)} créés, {len(errs)} errs")
                    if errs:
                        print(f"    Sample : {errs[0]}")

            # Fetch existing (session_id, trainer_id) pour dedup persistant
            existing_ft = rest_get("formation_trainers", select="session_id,trainer_id,loris_external_id")
            existing_pairs = {(r["session_id"], r["trainer_id"]) for r in (existing_ft if isinstance(existing_ft, list) else [])}
            existing_ext_ft = {r["loris_external_id"] for r in (existing_ft if isinstance(existing_ft, list) else []) if r.get("loris_external_id")}
            print(f"  ℹ️  {len(existing_pairs)} (session,trainer) pairs déjà en DB")

            to_insert, skipped_match, skipped_dup, skipped_existing = [], 0, 0, 0
            seen_ext = set()
            seen_pairs = set()
            for row in data:
                p = map_formation_trainer(row, lk["sessions_by_title"], lk["trainers_by_name"], lk.get("code_to_session"))
                if not p:
                    continue
                if "_skip_reason" in p:
                    skipped_match += 1
                    continue
                pair = (p["session_id"], p["trainer_id"])
                ext = p["loris_external_id"]
                if ext in seen_ext or ext in existing_ext_ft:
                    skipped_dup += 1
                    continue
                if pair in seen_pairs or pair in existing_pairs:
                    skipped_existing += 1
                    continue
                seen_ext.add(ext)
                seen_pairs.add(pair)
                to_insert.append(p)
            print(f"  → {len(to_insert)} à insérer, {skipped_match} no-match, {skipped_dup} doublons loris, {skipped_existing} pairs (session,trainer) déjà liées")
            inserted, errors = insert_batch("formation_trainers", to_insert, dry, ignore_duplicates=True)
            print(f"  ✅ inserted={len(inserted)} | errors={len(errors)}")
            if errors:
                print(f"  Sample : {errors[0]}")
            report["tables"]["formation_trainers"] = {
                "total_rows": len(data),
                "to_insert": len(to_insert),
                "skipped_no_match": skipped_match,
                "skipped_duplicates": skipped_dup,
                "inserted": len(inserted),
                "errors": len(errors),
                "error_samples": errors[:3],
            }

    # 6. ENROLLMENTS
    if "enrollments" in selected:
        print(f"\n{'─'*70}\n📂 ENROLLMENTS\n{'─'*70}")
        headers, data = read_xlsx(FILES["enrollments"])
        if not data:
            print(f"  ⚠️  {FILES['enrollments']} introuvable")
        else:
            # Bootstrap : créer les learners Loris manquants (référencés dans Suivi stagiaires mais absents de Apprenants.xlsx)
            referenced_names = {norm(r.get("Nom")) for r in data if norm(r.get("Nom"))}
            missing_learners = [n for n in referenced_names if norm_name(n) not in lk["learners_by_name"]]
            if missing_learners:
                print(f"  🔧 Bootstrap : {len(missing_learners)} learners Loris manquants à créer")
                bootstrap_payloads = []
                seen_bs = set()
                for name in missing_learners:
                    first, last = split_name(name)
                    ext = stable_external_id("learner", first or "", last or "", "")
                    if ext in seen_bs or ext in lk["learners_by_loris_id"]:
                        continue
                    seen_bs.add(ext)
                    bootstrap_payloads.append({
                        "entity_id": MR_ENTITY_ID,
                        "first_name": first or "—",
                        "last_name": last or name,
                        "loris_external_id": ext,
                        "loris_metadata": {"_bootstrap_from_enrollments": True},
                    })
                if bootstrap_payloads and not dry:
                    bs_inserted, bs_errs = insert_batch("learners", bootstrap_payloads, dry)
                    for r in bs_inserted:
                        fn = r.get("first_name") or ""
                        ln = r.get("last_name") or ""
                        for v in [f"{ln} {fn}", f"{fn} {ln}", ln, fn]:
                            k = norm_name(v)
                            if k:
                                lk["learners_by_name"][k] = r["id"]
                    print(f"  ✅ Bootstrap learners : {len(bs_inserted)} créés, {len(bs_errs)} errs")
                    if bs_errs:
                        print(f"    Sample : {bs_errs[0]}")

            # Fetch existing loris_external_id pour dedup persistant
            existing = rest_get("enrollments", select="loris_external_id",
                                **{"loris_external_id": "not.is.null"})
            existing_ext = {r["loris_external_id"] for r in (existing if isinstance(existing, list) else []) if r.get("loris_external_id")}
            print(f"  ℹ️  {len(existing_ext)} enrollments loris-tagged déjà en DB")

            to_insert, skipped_match, skipped_dup = [], 0, 0
            seen_ext = set()
            for row in data:
                p = map_enrollment(row, lk["sessions_by_title"], lk["learners_by_name"], lk.get("code_to_session"))
                if not p:
                    continue
                if "_skip_reason" in p:
                    skipped_match += 1
                    continue
                ext = p["loris_external_id"]
                if ext in seen_ext or ext in existing_ext:
                    skipped_dup += 1
                    continue
                seen_ext.add(ext)
                to_insert.append(p)
            print(f"  → {len(to_insert)} à insérer, {skipped_match} skippés (no match session/learner), {skipped_dup} doublons")
            inserted, errors = insert_batch("enrollments", to_insert, dry)
            print(f"  ✅ inserted={len(inserted)} | errors={len(errors)}")
            if errors:
                print(f"  Sample : {errors[0]}")
            report["tables"]["enrollments"] = {
                "total_rows": len(data),
                "to_insert": len(to_insert),
                "skipped_no_match": skipped_match,
                "skipped_duplicates": skipped_dup,
                "inserted": len(inserted),
                "errors": len(errors),
                "error_samples": errors[:3],
            }

    # 7. CRM_QUOTES
    if "crm_quotes" in selected:
        print(f"\n{'─'*70}\n📂 CRM_QUOTES\n{'─'*70}")
        headers, data = read_xlsx(FILES["crm_quotes"])
        if not data:
            print(f"  ⚠️  {FILES['crm_quotes']} introuvable")
        else:
            to_insert = []
            for idx, row in enumerate(data):
                p = map_crm_quote(row, idx, lk["clients_by_name"])
                if p:
                    to_insert.append(p)
            print(f"  → {len(to_insert)} devis à insérer")
            inserted, errors = insert_batch("crm_quotes", to_insert, dry)
            print(f"  ✅ inserted={len(inserted)} | errors={len(errors)}")
            if errors:
                print(f"  Sample : {errors[0]}")
            report["tables"]["crm_quotes"] = {
                "total_rows": len(data),
                "to_insert": len(to_insert),
                "inserted": len(inserted),
                "errors": len(errors),
                "error_samples": errors[:3],
            }

    # 8. FORMATION_INVOICES
    if "formation_invoices" in selected:
        print(f"\n{'─'*70}\n📂 FORMATION_INVOICES\n{'─'*70}")
        headers, data = read_xlsx(FILES["formation_invoices"])
        if not data:
            print(f"  ⚠️  {FILES['formation_invoices']} introuvable")
        else:
            # Récupère le MAX(global_number) existant pour éviter collision
            existing = rest_get("formation_invoices",
                                select="global_number",
                                order="global_number.desc.nullslast",
                                limit="1",
                                **{"entity_id": f"eq.{MR_ENTITY_ID}"})
            max_gn = 0
            if isinstance(existing, list) and existing and existing[0].get("global_number"):
                max_gn = existing[0]["global_number"]
            offset_gn = max(max_gn + 1, 900000)
            print(f"  ℹ️  Offset global_number : {offset_gn} (max existant = {max_gn})")

            # Bootstrap : créer les recipients manquants (financeurs/OPCO/etc.) en tant que clients
            referenced_recipients = set()
            for row in data:
                # Lignes Type « charge »-like (exactes → formation_charges, non exactes →
                # skippées entièrement) : leurs destinataires (formateurs/fournisseurs) ne
                # doivent PAS générer de fiches clients bootstrap.
                if "charge" in (norm(row.get("Type")) or "").lower():
                    continue
                rec = norm(row.get("Destinataire")) or norm(row.get("Client"))
                if rec:
                    referenced_recipients.add(rec)
            missing_recipients = [r for r in referenced_recipients if norm_name(r) not in lk["clients_by_name"]]
            if missing_recipients:
                print(f"  🔧 Bootstrap : {len(missing_recipients)} recipients (financeurs/OPCO) à créer comme clients")
                bs_payloads = []
                seen_bs = set()
                for name in missing_recipients:
                    ext = stable_external_id("client", name, "")
                    if ext in seen_bs or ext in lk["clients_by_loris_id"]:
                        continue
                    seen_bs.add(ext)
                    bs_payloads.append({
                        "entity_id": MR_ENTITY_ID,
                        "company_name": name,
                        "loris_external_id": ext,
                        "loris_metadata": {"_bootstrap_from_invoices": True, "_role": "financeur_ou_opco"},
                    })
                if bs_payloads and not dry:
                    bs_inserted, bs_errs = insert_batch("clients", bs_payloads, dry)
                    for r in bs_inserted:
                        if r.get("company_name"):
                            lk["clients_by_name"][norm_name(r["company_name"])] = r["id"]
                        if r.get("loris_external_id"):
                            lk["clients_by_loris_id"][r["loris_external_id"]] = r["id"]
                    print(f"  ✅ Bootstrap clients : {len(bs_inserted)} créés, {len(bs_errs)} errs")
                    if bs_errs:
                        print(f"    Sample : {bs_errs[0]}")

            # Set existing references to skip duplicates by external_reference
            existing_refs = rest_get("formation_invoices",
                                     select="external_reference",
                                     **{"entity_id": f"eq.{MR_ENTITY_ID}",
                                        "external_source": "eq.loris"})
            ref_set = {r["external_reference"] for r in (existing_refs if isinstance(existing_refs, list) else []) if r.get("external_reference")}

            to_insert, skipped_match, skipped_dup = [], 0, 0
            charge_candidates = []   # (base_dédupe, payload formation_charges) — lignes Type='charge'
            charge_skips = []        # raisons de skip côté charges (no-match session, montant illisible)
            chargelike_types = {}    # Type contenant « charge » SANS égalité exacte → ligne skippée ENTIÈREMENT
            unknown_types = {}       # Type non reconnu (hors facture/avoir/acompte/vide) → compté, routage facture inchangé
            empty_types = 0          # Types vides — comptés en info, routage facture par défaut
            for idx, row in enumerate(data):
                typ = (norm(row.get("Type")) or "").lower()
                if typ == "charge":
                    # Ligne Charge (coût formateur/fournisseur) → formation_charges,
                    # JAMAIS formation_invoices (cf. reclassement des 220 lignes historiques).
                    cp = map_formation_charge(row, lk["sessions_by_title"], lk.get("code_to_session"))
                    if "_skip_reason" in cp:
                        charge_skips.append(cp["_skip_reason"])
                        continue
                    raw_amount = to_decimal(row.get("Montant"))
                    if raw_amount is not None and raw_amount >= 0:
                        print(f"  ⚠️  Charge à montant ≥ 0 dans la source : « {cp['label']} » "
                              f"({raw_amount}) — anomalie source, importée avec abs()")
                    charge_candidates.append((cp.pop("_dedupe_base"), cp))
                    continue
                if "charge" in typ:
                    # Type « charge »-like sans égalité EXACTE (ex. « Charges ») : ligne skippée
                    # ENTIÈREMENT (ni facture ni charge) — jamais de facture parasite recréée en silence.
                    chargelike_types[typ] = chargelike_types.get(typ, 0) + 1
                    continue
                if typ == "":
                    empty_types += 1
                elif typ not in ("facture", "avoir", "acompte"):
                    # Type inconnu : compté et tracé (routage facture inchangé).
                    unknown_types[typ] = unknown_types.get(typ, 0) + 1
                p = map_formation_invoice(row, idx, lk["sessions_by_title"], lk["clients_by_name"], lk.get("code_to_session"))
                if not p:
                    continue
                if "_skip_reason" in p:
                    skipped_match += 1
                    continue
                if p.get("external_reference") and p["external_reference"] in ref_set:
                    skipped_dup += 1
                    continue
                # Override number + global_number avec offset unique
                p["number"] = offset_gn + idx
                p["global_number"] = offset_gn + idx
                to_insert.append(p)
            if chargelike_types:
                for t in sorted(chargelike_types):
                    print(f"  ⚠️  Type « charge »-like non exact : '{t}' — {chargelike_types[t]} ligne(s) "
                          f"skippée(s) ENTIÈREMENT (ni facture ni charge)")
            if unknown_types:
                print(f"  ⚠️  {sum(unknown_types.values())} ligne(s) de Type non reconnu "
                      f"(hors facture/avoir/acompte/charge) — traitées comme factures :")
                for t in sorted(unknown_types):
                    print(f"      • Type='{t}' : {unknown_types[t]} ligne(s)")
            if empty_types:
                print(f"  ℹ️  {empty_types} ligne(s) sans Type — routées en facture (défaut historique)")
            print(f"  → {len(to_insert)} factures à insérer, {skipped_match} skippées (no match), {skipped_dup} doublons (déjà importé)")
            inserted, errors = insert_batch("formation_invoices", to_insert, dry)
            print(f"  ✅ inserted={len(inserted)} | errors={len(errors)}")
            if errors:
                print(f"  Sample : {errors[0]}")
            report["tables"]["formation_invoices"] = {
                "total_rows": len(data),
                "to_insert": len(to_insert),
                "skipped_no_match": skipped_match,
                "skipped_duplicates": skipped_dup,
                "skipped_chargelike_entier": sum(chargelike_types.values()),
                "chargelike_types": chargelike_types,
                "unknown_types": unknown_types,
                "empty_types": empty_types,
                "inserted": len(inserted),
                "errors": len(errors),
                "error_samples": errors[:3],
            }

            # ── CHARGES (Type='charge') → formation_charges ──────────────────
            print(f"\n  📎 CHARGES (Type='charge') → formation_charges : "
                  f"{len(charge_candidates) + len(charge_skips)} ligne(s) source")
            if charge_skips:
                print(f"  ⏭️  {len(charge_skips)} charge(s) skippée(s) (no-match session / montant illisible) — échantillon :")
                for reason in charge_skips[:5]:
                    print(f"      • {reason}")

            # VERROU D'ORDRE (Design Notes v3) : tant que des factures parasites
            # « Loris Charge — » subsistent pour l'entité, AUCUNE insertion de charge
            # (sinon import et reclassement écriraient formation_charges en concurrence).
            # Le routage des lignes Type='charge' HORS factures reste actif ci-dessus.
            # soft=True : ce GET tourne APRÈS l'insertion des factures — une erreur bloque
            # les charges (conservateur) mais laisse le run écrire son rapport JSON.
            # Filtre client startswith : MÊME prédicat que le reclassement (l'ILIKE est
            # insensible à la casse) — périmètres alignés, pas de blocage sans issue.
            _legacy_rows = rest_get_all(
                "formation_invoices",
                soft=True,
                select="id,notes",
                **{"entity_id": f"eq.{MR_ENTITY_ID}",
                   "external_source": "eq.loris",
                   "notes": f"ilike.{CHARGE_NOTES_PREFIX}*"})
            legacy_charge_invoices = (None if _legacy_rows is None else
                                      [r for r in _legacy_rows
                                       if (r.get("notes") or "").startswith(CHARGE_NOTES_PREFIX)])
            if legacy_charge_invoices is None:
                print(f"  ⛔ GET du verrou en erreur — blocage CONSERVATEUR : les "
                      f"{len(charge_candidates)} insertion(s) de charges sont skippées "
                      "(relancer l'import une fois le réseau/API rétabli).")
                report["tables"]["formation_charges"] = {
                    "charge_rows": len(charge_candidates) + len(charge_skips),
                    "charges_blocked": len(charge_candidates),
                    "blocked_reason": "GET verrou en erreur (blocage conservateur)",
                    "skipped_no_match_ou_illisible": len(charge_skips),
                    "to_insert": 0,
                    "skipped_duplicates": 0,
                    "inserted": 0,
                    "errors": 0,
                    "error_samples": [],
                }
            elif legacy_charge_invoices:
                print(f"  ⛔ {len(legacy_charge_invoices)} facture(s) « {CHARGE_NOTES_PREFIX.strip()} » "
                      f"encore dans formation_invoices — TOUTES les insertions de charges sont "
                      f"skippées ({len(charge_candidates)} candidate(s)) :")
                print("     jouer scripts/import-loris/reclass_loris_charges.py d'abord, puis relancer cet import.")
                print("     ⚠️  Au re-run post-reclassement : pour un xlsx inchangé, le compteur « charges à")
                print("        insérer » doit rester 0 — un compte > 0 peut signaler des sessions re-résolues")
                print("        depuis l'import d'origine (code/titre) : vérifier en dry-run avant --execute.")
                report["tables"]["formation_charges"] = {
                    "charge_rows": len(charge_candidates) + len(charge_skips),
                    "charges_blocked": len(charge_candidates),
                    "skipped_no_match_ou_illisible": len(charge_skips),
                    "to_insert": 0,
                    "skipped_duplicates": 0,
                    "inserted": 0,
                    "errors": 0,
                    "error_samples": [],
                }
            else:
                # Dédupe MULTISET par (entity_id, session_id, base, montant quantizé) :
                # base candidate = norm_name(Destinataire) ; base existante = norm_name(label
                # sans préfixe ni suffixe ' (...)' FINAL) — cf. charge_label_base. Pour chaque
                # clé, n'insérer que max(0, n_candidats − n_existants) : tolère les vraies
                # charges multiples identiques (pas de drop silencieux d'une charge légitime)
                # et un re-run import→import donne 0 insertion.
                # soft=True : jamais de dédupe sur un GET échoué, mais ce GET tourne APRÈS
                # l'insertion des factures — une erreur bloque les charges (conservateur,
                # candidats vidés → 0 insertion) sans priver l'opérateur du rapport JSON.
                existing_charges = rest_get_all(
                    "formation_charges",
                    soft=True,
                    select="session_id,label,amount",
                    **{"entity_id": f"eq.{MR_ENTITY_ID}",
                       "label": f"like.{CHARGE_LABEL_PREFIX}*"})
                charges_blocked_on_error = 0
                if existing_charges is None:
                    print(f"  ⛔ GET de dédupe en erreur — blocage CONSERVATEUR : les "
                          f"{len(charge_candidates)} insertion(s) de charges sont skippées "
                          "(relancer l'import une fois le réseau/API rétabli).")
                    charges_blocked_on_error = len(charge_candidates)
                    charge_candidates = []
                    existing_charges = []
                existing_counts = {}
                for r in existing_charges:
                    key = (MR_ENTITY_ID, r.get("session_id"),
                           charge_label_base(r.get("label")),
                           canon_charge_amount(r.get("amount")))
                    existing_counts[key] = existing_counts.get(key, 0) + 1
                charges_to_insert, charges_dup_samples, charges_skipped_dup = [], [], 0
                remaining_existing = dict(existing_counts)
                for base, cp in charge_candidates:
                    key = (MR_ENTITY_ID, cp["session_id"], base, canon_charge_amount(cp["amount"]))
                    if remaining_existing.get(key, 0) > 0:
                        # Une charge existante « consomme » ce candidat (comptage multiset).
                        remaining_existing[key] -= 1
                        charges_skipped_dup += 1
                        if len(charges_dup_samples) < 5:
                            charges_dup_samples.append(f"{cp['label']} — {cp['amount']}")
                        continue
                    charges_to_insert.append(cp)
                if charges_skipped_dup:
                    print(f"  ⏭️  {charges_skipped_dup} candidat(s) excédentaire(s) droppé(s) "
                          f"(dédupe multiset) — échantillon :")
                    for sample in charges_dup_samples:
                        print(f"      • {sample}")
                print(f"  → {len(charges_to_insert)} charges à insérer")
                ch_inserted, ch_errors = insert_batch("formation_charges", charges_to_insert, dry)
                print(f"  ✅ charges inserted={len(ch_inserted)} | errors={len(ch_errors)}")
                if ch_errors:
                    print(f"  Sample : {ch_errors[0]}")
                report["tables"]["formation_charges"] = {
                    "charge_rows": len(charge_candidates) + len(charge_skips) + charges_blocked_on_error,
                    "charges_blocked": charges_blocked_on_error,
                    "skipped_no_match_ou_illisible": len(charge_skips),
                    "to_insert": len(charges_to_insert),
                    "skipped_duplicates": charges_skipped_dup,
                    "inserted": len(ch_inserted),
                    "errors": len(ch_errors),
                    "error_samples": ch_errors[:3],
                }

    # Final report
    report["ended_at"] = datetime.utcnow().isoformat() + "Z"
    REPORT_PATH.write_text(json.dumps(report, indent=2, ensure_ascii=False, default=str))
    print(f"\n{'═'*70}\n📊 Rapport écrit : {REPORT_PATH}\n{'═'*70}")


if __name__ == "__main__":
    main()
