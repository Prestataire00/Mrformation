#!/usr/bin/env python3
"""
Reconstruit les CONTACTS clients depuis Clients.xlsx (l'import initial ne les a pas créés :
4 contacts en base / 267 clients, alors que la source contient « Représenté par » + Email + Tel).

Source par client (Clients.xlsx) :
  - Contact principal : « Représenté par » (nom) + « Email » + « Tel » + « Fonction du représentant ».
  - Contacts secondaires : « Contact 2..5 » + « Email 2..5 ».

Règles :
  - Match Nom (Clients.xlsx) → client DB par nom normalisé (fiche canonique = min id si doublons).
  - Idempotent : on SAUTE tout client ayant déjà ≥1 contact.
  - first_name/last_name sont NOT NULL : on parse « NOM Prénom » ; si le nom est inexploitable
    (vide / 1 lettre) mais qu'un email réel existe → contact dérivé du local-part de l'email.
  - Email générique (placeholder @mrformation.fr / @c3vformation.fr) → email laissé NULL sur le contact.
  - Si ni nom exploitable ni email réel → pas de contact (skip).

DRY-RUN par défaut. `--apply` pour écrire.
  python3 scripts/import-loris/apply_client_contacts.py [--apply]
"""

import argparse
import importlib.util
import json
import re
import urllib.request
from collections import defaultdict
from pathlib import Path

_spec = importlib.util.spec_from_file_location("rec", Path(__file__).parent / "reconcile_code_formation.py")
rec = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(rec)

REPORT_PATH = Path(__file__).parent / "apply_client_contacts_report.json"
GENERIC_DOMAINS = ("mrformation.fr", "c3vformation.fr", "c3v-formation.fr")


def split_name(full):
    """'NOM Prénom' (NOM en majuscules) → (first, last). 1 token → ('—', token)."""
    full = (full or "").strip()
    parts = full.split()
    if not parts:
        return None, None
    if len(parts) == 1:
        return "—", parts[0]
    if parts[0].isupper() and len(parts[0]) > 1:
        return " ".join(parts[1:]), parts[0]
    return parts[0], " ".join(parts[1:])


def name_usable(rep):
    """Un nom est exploitable s'il a au moins un token alphabétique de longueur ≥ 2."""
    if not rep:
        return False
    return any(len(t) >= 2 and any(ch.isalpha() for ch in t) for t in rep.split())


def is_generic_email(email):
    e = (email or "").lower()
    return any(e.endswith("@" + d) for d in GENERIC_DOMAINS)


def email_localpart(email):
    return (email or "").split("@")[0]


def rest_post(table, rows):
    req = urllib.request.Request(
        f"{rec.SUPABASE_URL}/rest/v1/{table}", data=json.dumps(rows).encode(),
        headers={"apikey": rec.SERVICE_ROLE, "Authorization": f"Bearer {rec.SERVICE_ROLE}",
                 "Content-Type": "application/json", "Prefer": "return=minimal"},
        method="POST")
    with urllib.request.urlopen(req) as r:
        return r.status


def read_clients_xlsx():
    from openpyxl import load_workbook
    path = rec.DOWNLOADS / "Clients.xlsx"
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    out = []
    for r in rows[1:]:
        d = {hdr[i]: v for i, v in enumerate(r) if i < len(hdr)}
        if rec.norm(d.get("Nom")):
            out.append(d)
    return out


def build_contact(name_raw, email_raw, phone_raw=None, job_raw=None, primary=False):
    """Retourne un dict contact partiel (sans client_id) ou None si pas d'identité exploitable."""
    name = rec.norm(name_raw)
    email = rec.norm(email_raw)
    generic = is_generic_email(email)
    if name_usable(name):
        first, last = split_name(name)
    elif email and not generic:
        first, last = "—", email_localpart(email).upper()
    else:
        return None
    phone = rec.norm(phone_raw)
    if phone and sum(ch.isdigit() for ch in phone) < 6:  # '0', '-', junk → pas un vrai numéro
        phone = None
    return {
        "first_name": first or "—",
        "last_name": last or "—",
        "email": None if (not email or generic) else email.lower(),
        "phone": phone,
        "job_title": rec.norm(job_raw),
        "is_primary": primary,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()
    print(f"CONTACTS clients — {'APPLY' if args.apply else 'DRY-RUN'} — MR FORMATION\n")

    db_clients = rec.rest_get_all("clients", select="id,company_name", entity_id=f"eq.{rec.MR_ENTITY_ID}")
    existing = rec.rest_get_all("contacts", select="client_id")
    has_contact = {x["client_id"] for x in existing}
    client_by_name = defaultdict(list)
    for c in db_clients:
        client_by_name[rec.norm_name(c.get("company_name"))].append(c["id"])

    rows = read_clients_xlsx()
    to_insert = []
    stats = {"clients_source": len(rows), "non_matche": 0, "deja_contact": 0,
             "primary_rep": 0, "primary_email": 0, "secondaires": 0, "sans_identite": 0}
    unmatched = []
    for r in rows:
        cids = client_by_name.get(rec.norm_name(r.get("Nom")), [])
        if not cids:
            stats["non_matche"] += 1
            if len(unmatched) < 20:
                unmatched.append(rec.norm(r.get("Nom")))
            continue
        client_id = sorted(cids)[0]  # fiche canonique si doublons
        if client_id in has_contact:
            stats["deja_contact"] += 1
            continue

        client_contacts = []
        # contact principal
        primary = build_contact(r.get("Représenté par"), r.get("Email"),
                                 r.get("Tel"), r.get("Fonction du représentant"), primary=True)
        if primary:
            client_contacts.append(primary)
            if name_usable(rec.norm(r.get("Représenté par"))):
                stats["primary_rep"] += 1
            else:
                stats["primary_email"] += 1
        else:
            stats["sans_identite"] += 1
        # contacts secondaires 2..5
        for n in range(2, 6):
            sec = build_contact(r.get(f"Contact {n}"), r.get(f"Email {n}"), primary=False)
            if sec:
                client_contacts.append(sec)
                stats["secondaires"] += 1

        for c in client_contacts:
            to_insert.append({"client_id": client_id, **c})

    report = {
        "mode": "apply" if args.apply else "dry-run",
        "contacts_a_creer": len(to_insert),
        "clients_couverts": len({c["client_id"] for c in to_insert}),
        "stats": stats,
        "clients_source_non_matches": unmatched,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2))

    print(f"Contacts à créer : {len(to_insert)} (sur {report['clients_couverts']} clients)")
    print(f"  principal via 'Représenté par' : {stats['primary_rep']}")
    print(f"  principal via email (nom inexploitable) : {stats['primary_email']}")
    print(f"  secondaires (Contact 2..5) : {stats['secondaires']}")
    print(f"Sautés : déjà un contact={stats['deja_contact']}, sans identité exploitable={stats['sans_identite']}, client non matché={stats['non_matche']}")
    if unmatched:
        print("  clients source non matchés (échantillon) :", unmatched[:8])

    if not args.apply:
        print(f"\n[DRY-RUN] Aucune écriture. Rapport : {REPORT_PATH}")
        return

    print("\n>>> APPLICATION…")
    done = 0
    for batch in [to_insert[i:i + 200] for i in range(0, len(to_insert), 200)]:
        rest_post("contacts", batch)
        done += len(batch)
    print(f"✅ {done} contacts créés. Rapport : {REPORT_PATH}")


if __name__ == "__main__":
    main()
